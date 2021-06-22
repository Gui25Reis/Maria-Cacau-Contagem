import matplotlib.pyplot as plt, pandas as pd, time as tm

# 'G:/Users/guui_/Desktop/GuiTestes/Pastas/Maria Cacau/Arquivos/Arq MC.xlsx'
local_arq = 'G:/OneDrive/Gui/GuiTestes/Python/Maria Cacau/Arquivos/Arq MC.xlsx' 
arq = pd.read_excel(local_arq, decimal='.')

# 'G:/Users/guui_/Desktop/GuiTestes/Pastas/Maria Cacau/Arquivos/Resumo pedidos.txt'
local_txt = 'G:/OneDrive/Gui/GuiTestes/Python/Maria Cacau/Arquivos/Resumo pedidos.txt'#'Arquivos/Resumo pedidos.txt'
arq_peds = open(local_txt,'w')

from openpyxl import *
arq_2 = load_workbook(local_arq)
plan = arq_2['Nota - SAGE']


def salv_foto(dia_, num_):
    plt.rc('savefig',pad_inches=0.5,bbox='tight')
    # 'C:/Users/guui_/Desktop/GuiTestes/Pastas/Maria Cacau/Arquivos/Fotos/Moto {} P{}.png'
    plt.savefig('G:/OneDrive/Gui/GuiTestes/Python/Maria Cacau/Arquivos/Fotos/Moto {} P{}.png' .format(dia_, num_))


def pedidos(dia_):
    filt = arq.loc[(arq['DATA ENTREGA'] == '{} 00:00:00' .format(dia_))] [cols] 
    filt['ind'] = [x for x in range(len(filt))]
    filt = pd.DataFrame(filt) .set_index("ind")

    quant = filt['Modal'].value_counts()
    tipo = quant.index.tolist()

    entregas = ''
    for x in range (len(quant)):
        entregas += '\n' + '{} = {}' .format(tipo[x],quant[x])

    pag = filt[cols[1]]

    dev = ''
    for x in range(len(pag)):
        if pag[x] < 0 and filt[cols[3]][x] != 'FABRICA':
            dev += '\n' + '{} -> {} | {} | {} | $: {} \n' .format(
            filt[cols[0]][x], filt[cols[4]][x], filt[cols[3]][x], filt[cols[-3]][x], filt[cols[1]][x])

    return '''Para o dia {} temos: {} pedidos \n{} \n\nFalta(m) pagar: \n{}''' .format(dia[:5], sum(quant), entregas, dev)


def endereco (dia_):
    filt = arq.loc[(arq['DATA ENTREGA'] == '{} 00:00:00' .format(dia_)) & (arq['Modal'] == 'MOTOBOY')] [cols] 
    filt['ind'] = [x for x in range(len(filt))]
    filt = pd.DataFrame(filt) .set_index("ind")

    for x in range(len(filt)):
        cond = ['Complemento:','Compl.']
        cc = filt[cols[6]][x]
        try:
            for y in cond:
                if len(filt[cols[6]][x].split(':')) != 1:
                    if y.upper() in cc.upper():
                        comp = filt[cols[6]][x].split(':')[1]
                        break
                else:
                    comp = cc
        except:
            comp == cc

            
    
        fig = plt.figure()
        plt.rc('figure',figsize= (18, 11.0))
        ax = fig.add_subplot(111)
        fig.subplots_adjust(top=0.85)
        plt.xticks([])
        plt.yticks([])

        dest = '                           DESTINATÁRIO:                         '
        ax.text(0.015, 0.9   , dest,              fontsize=40,bbox={'facecolor':'yellow', 'alpha':1, 'pad':12})

        ax.text(0, 0.8   , 'Nome: {}'        .format( filt[cols[4]][x]  ),         fontsize=30)
        ax.text(0, 0.725 , 'End: {}'         .format( filt[cols[5]][x]  ),         fontsize=30)
        ax.text(0, 0.645 , 'Complemento: {}' .format( comp ),                      fontsize=30)
        ax.text(0, 0.565 , 'Bairro: {}'      .format( filt[cols[7]][x]  ),         fontsize=30)
        ax.text(0, 0.485 , 'Cidade: {}'      .format( filt[cols[8]][x].split('-')[0] ),         fontsize=30)
        ax.text(0, 0.41  , 'Estado: {}'      .format( filt[cols[8]][x].split('-')[1].upper() ), fontsize=30)
        ax.text(0, 0.335 , 'Cep: {}'         .format( filt[cols[9]][x]  ),         fontsize=30)
        ax.text(0, 0.255 , 'Tel: {}'         .format( filt[cols[10]][x] ),         fontsize=30)
        ax.text(0.85, 0.15 , 'P{}'           .format( filt[cols[0]][x]  ),         fontsize=30)

        rem = '''                                                    Remetente: Maria Cacau
        Rua Amácio Mazzaropi, 113 - Jd Palermo - Nova Petrópolis - São Bernardo do Campo - SP'''
        rem2 ='''                                           CEP 09780-480 - Tel: 11-98081-5558'''
        ax.text(0.17, 0.07, rem,              fontsize=15)
        ax.text(0.17, 0.033, rem2,            fontsize=15)

        salv_foto(dia_,filt[cols[0]][x])


def nota_SAGE (dia_):
    #dia_ = dia_form
    cols = [arq_cols[x] for x in range(15,43,4)]
    prods = arq.loc[(arq['DATA ENTREGA'] == '{} 00:00:00' .format(dia_)) & (arq['Modal'] != 'MOTOBOY') & (arq['Evento'] != 'Amostras')] [cols] 
    prods['ind'] = [x for x in range(len(prods))]
    prods = pd.DataFrame(prods) .set_index("ind")

    veri = lambda x: 'Sim' if 'BELGA' in str(set(prods.loc[x])) else 'Não'
    belga = list(map(veri, range(len(prods))))

    
    filt = arq.loc[(arq['DATA ENTREGA'] == '{} 00:00:00' .format(dia_)) & (arq['Modal'] != 'MOTOBOY') & (arq['Evento'] != 'Amostras')] [cols_usad] 

    compl = []
    for y in list(filt[cols_usad[7]]):
        try: 
            if  'complemento:' in y.lower() or 'compl.' in y.lower():
                compl.append(y.split(':')[1])
            else:
                compl.append(y)
        except:
                compl.append(y)
        
   
    total = []
    tot = list(filt[cols_usad[11]])
    fret = list(filt[cols_usad[10]])
    for x in range(len(tot)):
        total.append(str(round(tot[x]-fret[x])))
    

    filt['Belga?'] = belga
    filt['ind'] = [x for x in range(len(filt))]
    filt = pd.DataFrame(filt) .set_index("ind")

    colunas = list(filt.columns)
    colunas.insert(12,'Belga?')
    colunas.insert(10,'')
    colunas.insert(14,'')
    
    modal_nota = ['SEDEX','SEDEX 10','SEDEX 12','PAC','ELO7 PAC','ELO7 CORREIO']

    num_coluna = 2
    for col in range(len(filt)):
        if filt.iloc[col][1] in modal_nota:
            for lin in range(len(colunas)-1):
                if lin == 7:
                    plan.cell(row=lin+1, column=num_coluna, value=str(compl[col]))
                elif lin == 11:
                    plan.cell(row=lin+1, column=num_coluna, value=str(fret[col]))
                elif lin == 12:
                    plan.cell(row=lin+1, column=num_coluna, value=str(total[col]))
                elif lin not in [10,14] and lin not in [7,11,12]:
                    plan.cell(row=lin+1, column=num_coluna, value=str(filt[colunas[lin]] [col]))
            num_coluna += 3         


def funcs(dia_):
    arq_peds.write(pedidos(dia_))
    endereco(dia_)
    endereco(dia_)
    nota_SAGE(dia_)

    arq_2.save(local_arq)
    arq_2.close()
    arq_peds.close()

    print('\nPronto!!')

arq_cols = list(arq.columns)

cols = ['PEDIDO']
for x in arq_cols[73:85]:
    cols.append(x)


cols_usad = ['PEDIDO','Modal','NOME COMPRADOR','CPF','EMAIL','CEP','Rua',
'Compl.','Bairro','Cidade-UF','$FRETE','TOTAL','TEL','NEGOCIAÇÃO','Evento']

cols_plan = ['PEDIDO','Entrega','Nome','CPF','EMAIL','CEP','Rua','Compl.','Bairro','Cidade','',
'Frete','Total','Belga?','','TEL','Loc Tel','Evento']



ops = input('''\n\t\t|\/| /\ /? | /\   ( /\ ( /\ |_| 
\n01 -> Nota Fiscal \n\nOps: ''')

while ops not in ['1']:
    ops = input('Dado inválido. \n\nEscolha uma opção: ')

if ops == '1':
    print('\nDigite a data - dd/mm/aaaa')

    dia = input('\nData: ')
    while len(dia) != 10:
        print('Entrada inválida. Digite a data como dd/mm/aaaa')
        dia = input('\nDia: ')

    y = dia.split('/')[::-1]
    dia_form = '{}-{}-{}' .format(y[0],y[1],y[2])

    funcs(dia_form)



print('O programa vai fechar em ', end = '')
for x in range (3,0,-1):
    print(end = '{}.. ' .format(x))
    tm.sleep(1)
print(' ', end = '\r')
